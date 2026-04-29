import argparse
import csv
from pathlib import Path

from openpyxl import Workbook


SHEETS = [
    ("TopPriority", "top100_priority.csv"),
    ("TopActivity", "top100_by_activity.csv"),
    ("AllChats", "all_chats.csv"),
    ("FullHistoryRaw", "full_history_raw.csv"),
]


def append_csv_to_sheet(csv_path: Path, ws) -> int:
    rows_written = 0
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
            rows_written += 1
    return rows_written


def build_workbook(results_dir: Path, output_file: Path) -> None:
    wb = Workbook(write_only=True)

    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Telegram Lead Generator", "FABRICBOT ECOSYSTEM"])
    summary_ws.append(["Results Folder", str(results_dir)])
    summary_ws.append([])
    summary_ws.append(["Sheet", "Source CSV", "Rows"])

    for sheet_name, csv_name in SHEETS:
        csv_path = results_dir / csv_name
        if not csv_path.exists():
            summary_ws.append([sheet_name, csv_name, "missing"])
            continue

        ws = wb.create_sheet(sheet_name)
        row_count = append_csv_to_sheet(csv_path, ws)
        summary_ws.append([sheet_name, csv_name, row_count])

    wb.save(output_file)


def main() -> int:
    parser = argparse.ArgumentParser(description="Build XLSX workbook from lead CSV files")
    parser.add_argument("--results-dir", required=True, help="Directory with generated CSV files")
    parser.add_argument(
        "--output",
        default="telegram_leads_bundle.xlsx",
        help="Output xlsx file path",
    )
    args = parser.parse_args()

    results_dir = Path(args.results_dir).resolve()
    output_file = Path(args.output)
    if not output_file.is_absolute():
        output_file = results_dir / output_file

    output_file.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(results_dir, output_file)
    print(f"Workbook created: {output_file}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
